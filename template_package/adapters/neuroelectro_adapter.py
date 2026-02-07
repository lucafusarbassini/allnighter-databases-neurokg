"""
NeuroElectro Adapter for BioCypher.

Loads NeuroElectro data from Excel files and generates:
- NeuronType nodes (94 neuron types with NeuroLex IDs)
- NeuronEphysProperty edges (neuron → electrophysiology measurements)
"""

import re
from pathlib import Path
from biocypher._logger import logger

try:
    import openpyxl
except ImportError:
    openpyxl = None


class NeuroElectroAdapter:
    def __init__(self, data_dir="template_package/data/neuroelectro"):
        self.data_dir = Path(data_dir)
        self.neuron_types = {}  # neuroelectro_id -> {name, neurolex_id, ...}
        self.measurements = []  # aggregated ephys measurements
        self._load_data()

    def _sanitize(self, text):
        if text is None:
            return ""
        text = str(text)
        text = re.sub(r'<[^>]+>', '', text)
        text = text.replace('"', '""')
        text = text.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
        return text.strip()

    def _load_data(self):
        """Load NeuroElectro Excel data."""
        if openpyxl is None:
            logger.warning("NeuroElectro: openpyxl not installed, skipping")
            return

        # Load neuron type descriptions
        desc_path = self.data_dir / 'data_download' / 'neuron_description.xlsx'
        if desc_path.exists():
            self._load_neuron_descriptions(desc_path)

        # Load neurophysiology data
        data_path = self.data_dir / 'data_download' / 'neurophysiology_data.xlsx'
        if data_path.exists():
            self._load_neurophysiology(data_path)

        logger.info(f"NeuroElectro: Loaded {len(self.neuron_types)} neuron types, "
                     f"{len(self.measurements)} measurements")

    def _load_neuron_descriptions(self, path):
        """Load neuron type descriptions."""
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                name = row_dict.get('Neuron Type', '')
                ne_id = row_dict.get('NeuroElectro ID', '')
                neurolex_id = row_dict.get('NeuroLex ID', '')
                criteria = row_dict.get('Defining Criteria', '')

                if name and ne_id:
                    self.neuron_types[str(ne_id)] = {
                        'name': self._sanitize(name),
                        'neuroelectro_id': str(ne_id),
                        'neurolex_id': self._sanitize(str(neurolex_id)) if neurolex_id else '',
                        'defining_criteria': self._sanitize(str(criteria)) if criteria else '',
                    }

            wb.close()
            logger.info(f"NeuroElectro: Loaded {len(self.neuron_types)} neuron type descriptions")
        except Exception as e:
            logger.warning(f"NeuroElectro: Error loading neuron descriptions: {e}")

    def _load_neurophysiology(self, path):
        """Load neurophysiology measurements."""
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

            # Define ephys property columns
            ephys_cols = [
                'CellCapacitance', 'InputResistance', 'RestingMembranePotential',
                'MembraneTimeConstant', 'SpikeAmplitude', 'SpikeHalfWidth',
                'SpikeThreshold', 'Rheobase', 'FiringFrequency', 'AhpDuration',
                'CellDiameter', 'SagRatio', 'SpikeOvershoot', 'AhpAmplitude',
                'FiSlope', 'SpontaneousFiringRate', 'FastAhpAmplitude',
                'SpikeWidth', 'AdaptationRatio', 'SpikePeak',
            ]

            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                neuron_type = self._sanitize(str(row_dict.get('NeuronType', '')))
                species = self._sanitize(str(row_dict.get('Species', '')))

                if not neuron_type:
                    continue

                # Find matching neuron type ID
                neuron_id = None
                for nid, ndata in self.neuron_types.items():
                    if ndata['name'] == neuron_type:
                        neuron_id = nid
                        break

                if not neuron_id:
                    # Create a new neuron type entry
                    neuron_id = neuron_type.replace(' ', '_')
                    if neuron_id not in self.neuron_types:
                        self.neuron_types[neuron_id] = {
                            'name': neuron_type,
                            'neuroelectro_id': neuron_id,
                            'neurolex_id': '',
                            'defining_criteria': '',
                        }

                # Collect ephys measurements
                for prop in ephys_cols:
                    if prop in row_dict:
                        val = row_dict[prop]
                        if val is not None and str(val) != 'nan' and str(val).strip():
                            try:
                                float_val = float(val)
                                self.measurements.append({
                                    'neuron_id': neuron_id,
                                    'property': prop,
                                    'value': float_val,
                                    'species': species,
                                })
                            except (ValueError, TypeError):
                                pass

            wb.close()
            logger.info(f"NeuroElectro: Loaded {len(self.measurements)} measurements")
        except Exception as e:
            logger.warning(f"NeuroElectro: Error loading neurophysiology data: {e}")

    def get_nodes(self):
        """
        Generate neuron type nodes.
        Yields: (id, label, properties)
        """
        logger.info("NeuroElectro: Generating nodes...")
        count = 0

        for nid, ndata in self.neuron_types.items():
            node_id = f"NE:{nid}"
            props = {
                'name': ndata['name'],
                'neuroelectro_id': ndata['neuroelectro_id'],
                'neurolex_id': ndata['neurolex_id'],
                'defining_criteria': ndata['defining_criteria'],
                'source': 'NeuroElectro',
            }
            yield (node_id, "NeuronType", props)
            count += 1

        logger.info(f"NeuroElectro: Generated {count} NeuronType nodes")

    def get_edges(self):
        """
        Generate neuron-to-ephys-property edges.
        Yields: (id, source, target, label, properties)
        """
        logger.info("NeuroElectro: Generating edges...")
        count = 0

        # Aggregate measurements by neuron type + property
        from collections import defaultdict
        aggregated = defaultdict(list)
        for m in self.measurements:
            key = (m['neuron_id'], m['property'])
            aggregated[key].append(m['value'])

        for (neuron_id, prop), values in aggregated.items():
            mean_val = sum(values) / len(values)
            min_val = min(values)
            max_val = max(values)

            yield (
                None,
                f"NE:{neuron_id}",
                f"EPHYS:{prop}",
                "NeuronEphysProperty",
                {
                    'property_name': prop,
                    'mean_value': round(mean_val, 4),
                    'min_value': round(min_val, 4),
                    'max_value': round(max_val, 4),
                    'num_observations': len(values),
                }
            )
            count += 1

        logger.info(f"NeuroElectro: Generated {count} NeuronEphysProperty edges")
